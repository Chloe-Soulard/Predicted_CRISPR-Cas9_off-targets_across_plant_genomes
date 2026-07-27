"""
Step 1c — Build the Excel form for manually curating species BLAST cannot resolve.

Collects every (gene, species) pair still without a hit after step1_fallback.py
and writes one pre-filled row per pair. The `Status` column says whether the gap
is real:

    confirmed-absent           a clean query found nothing — curate this one
    network-err                a query failed; re-run the fallback instead
    NOT-YET-ATTEMPTED          step1_fallback.py has not looked at it yet

Fill the CDS column (at minimum), then run `build_manual_overrides.py` to convert
the sheet into config/manual_cds_overrides.json, which step 2 reads automatically.

Output: manual_overrides_TEMPLATE.xlsx

Run from pipeline/:
    python make_manual_template.py
"""

from __future__ import annotations

import argparse
from pathlib import Path

import openpyxl
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from paths import (
    ROOT, STEP1_DIR, gene_names, gene_slug, load_json, load_species_genomes,
)

OUT_FILE = ROOT / "manual_overrides_TEMPLATE.xlsx"

# (header, column width, filled by the curator?, cell comment)
# The column order is the contract with build_manual_overrides.py — keep them in step.
COLUMNS = [
    ("Gene", 12, False, "Pre-filled — do not edit"),
    ("Species", 24, False, "Pre-filled — do not edit"),
    ("Genome to search", 26, False,
     "CRISPOR genome id from species_genomes.json — search this assembly (e.g. Phytozome)"),
    ("Status", 16, False, "Why this row is here — see the README sheet"),
    ("Protein Accession", 22, True,
     "Source id of the ortholog (e.g. an NCBI or Phytozome accession); "
     "write 'manual' if it has no accession"),
    ("Protein Sequence", 40, True,
     "The ortholog protein. Used to validate the CDS: it must translate to exactly this."),
    ("CDS (required)", 40, True,
     "Full spliced coding sequence (ATG ... stop). REQUIRED — drives the conservation analysis."),
    ("Genomic Sequence", 40, True,
     "Optional: the genomic gene region (exons + introns) in coding orientation."),
    ("Reason / Notes", 30, True,
     "Why this entry was curated (e.g. 'low-homology auto hit replaced'), source database, caveats"),
]

SEQUENCE_COLUMNS = {"Protein Sequence", "CDS (required)", "Genomic Sequence"}

STATUS_BY_DB = {
    "none_targeted":  "confirmed-absent",
    "error_targeted": "network-err (re-run the fallback)",
    "none":           "NOT-YET-ATTEMPTED (run fallback first)",
    "":               "NOT-YET-ATTEMPTED (run fallback first)",
}

README_LINES = [
    "HOW TO USE THIS FORM",
    "",
    "1. Each row is one gene in one species that BLAST could not find in NCBI.",
    "2. GREY columns are pre-filled (Gene, Species, Genome to search, Status). Do not edit.",
    "3. YELLOW columns are for you to fill in.",
    "",
    "REQUIRED:",
    "  - CDS              : the spliced coding sequence (ATG ... stop codon) of the ortholog.",
    "                       This is what puts the species into the conservation analysis.",
    "  - Protein Sequence : the ortholog protein. build_manual_overrides.py checks that the",
    "                       CDS translates to exactly this, which catches a mis-pasted or",
    "                       frame-shifted sequence before it reaches the analysis.",
    "",
    "OPTIONAL:",
    "  - Protein Accession : source id of the ortholog (e.g. a Phytozome locus), or 'manual'.",
    "  - Genomic Sequence  : the genomic gene region (exons + introns) in coding (5'->3')",
    "        orientation. Leave it blank and the species still appears in the conservation",
    "        results, just without a genomic sequence.",
    "  - Reason / Notes    : why this entry was curated; recorded in the output JSON.",
    "",
    "STATUS COLUMN:",
    "  confirmed-absent   : a clean BLAST query found nothing. These are the rows to fill.",
    "  network-err        : a query failed. Re-run step1_fallback.py rather than curating.",
    "  NOT-YET-ATTEMPTED  : step1_fallback.py has not queried this species yet.",
    "",
    "WHEN DONE: save this file, then run:  python build_manual_overrides.py",
    "It writes config/manual_cds_overrides.json, which step 2 reads automatically.",
]


def collect_missing() -> list[tuple[str, str, str, str]]:
    """(gene, species, CRISPOR genome, status) for every pair still without a hit."""
    genomes = load_species_genomes()
    rows = []
    for gene in gene_names():
        step1 = load_json(STEP1_DIR / f"{gene_slug(gene)}.json")
        for species, entry in step1.items():
            if entry.get("hits"):
                continue
            db = entry.get("db", "")
            rows.append((gene, species, genomes.get(species, "?"),
                         STATUS_BY_DB.get(db, db)))
    return rows


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__.strip().splitlines()[0])
    parser.add_argument("--out", type=Path, default=OUT_FILE, help="output form path")
    args = parser.parse_args()

    rows = collect_missing()
    if not rows:
        print("Every gene x species pair has a BLAST hit — no curation needed.")
        return

    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Manual overrides"

    header_fill = PatternFill("solid", fgColor="2F5496")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    curator_fill = PatternFill("solid", fgColor="FFF2CC")   # yellow: you fill this
    locked_fill  = PatternFill("solid", fgColor="E7E6E6")   # grey: pre-filled
    network_fill = PatternFill("solid", fgColor="FCE4D6")   # orange: network error

    sheet.merge_cells(f"A1:{get_column_letter(len(COLUMNS))}1")
    title = sheet["A1"]
    title.value = ("Manual ortholog entry form — fill the YELLOW columns. "
                   "CDS is required; the genomic sequence is optional.")
    title.font = Font(bold=True, size=12, color="1F3864")
    sheet.row_dimensions[1].height = 20

    for col, (name, width, _is_curator, comment) in enumerate(COLUMNS, start=1):
        cell = sheet.cell(row=2, column=col, value=name)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.comment = Comment(comment, "template")
        sheet.column_dimensions[get_column_letter(col)].width = width
    sheet.row_dimensions[2].height = 34

    for row_i, (gene, species, genome, status) in enumerate(rows, start=3):
        prefilled = {1: gene, 2: species, 3: genome, 4: status}
        is_network_error = status.startswith("network")
        for col, (name, _width, _is_curator, _comment) in enumerate(COLUMNS, start=1):
            cell = sheet.cell(row=row_i, column=col)
            if col in prefilled:
                cell.value = prefilled[col]
                cell.fill = network_fill if is_network_error else locked_fill
                cell.font = Font(size=10)
            else:
                cell.fill = curator_fill
                cell.font = Font(size=10, name="Courier New")
            cell.alignment = Alignment(vertical="top", wrap_text=name in SEQUENCE_COLUMNS)
    sheet.freeze_panes = "A3"

    readme = workbook.create_sheet("README")
    for row_i, line in enumerate(README_LINES, start=1):
        cell = readme.cell(row=row_i, column=1, value=line)
        if row_i == 1:
            cell.font = Font(bold=True, size=13, color="1F3864")
        elif line.endswith(":") or (line and line.isupper()):
            cell.font = Font(bold=True, size=10)
        else:
            cell.font = Font(size=10)
    readme.column_dimensions["A"].width = 100

    workbook.save(args.out)

    n_confirmed = sum(1 for r in rows if r[3] == "confirmed-absent")
    print(f"Wrote {args.out}")
    print(f"  {len(rows)} rows: {n_confirmed} confirmed-absent (fill these), "
          f"{len(rows) - n_confirmed} other")


if __name__ == "__main__":
    main()
