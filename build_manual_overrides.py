"""
Step 1d — Convert the filled Excel form into config/manual_cds_overrides.json.

Only rows with a CDS are written; blank rows are reported and ignored. Step 2
reads this file and uses a curated entry in place of an NCBI lookup.

Every entry whose row also carries a protein sequence is validated by checking
that its CDS translates to exactly that protein — the gate that catches a
mis-pasted, truncated or frame-shifted sequence before it reaches the analysis.
A failure aborts the write unless --force is given.

Columns are located by *header name*, not by position, so both generations of the
form are readable: the current template's separate "Protein Sequence" column, and
the older full template where the protein sequence sits under the "Protein
Accession" header with Strand / Genomic Accession / Exons columns in between.
A cell under an accession header that is really a peptide is detected and stored
as the protein, so the translation gate still runs.

Gene labels are normalised against the panel in config/gene_proteins.json,
case-insensitively and via LEGACY_GENE_NAMES for forms filled before the panel
was renamed. An unrecognised label is an error, never a silently dropped row —
step 2 looks overrides up by the current panel name.

Output schema (gene -> species -> entry):
    {"ACT1": {"Rubus occidentalis": {
        "protein_accession": "manual",          # source id, or 'manual' if none
        "protein": "MADGE...",                  # validated against the CDS
        "cds": "ATG...",
        "cds_length": 1338,
        "manual_genomic_sequence": "ATG...",    # only if a genomic sequence was given
        "manual_modif_reason": "...",           # only if a reason was given
        "source": "manual"}}}

The conservation analysis needs only the CDS. No NCBI accession or strand is
required — these species are precisely the ones NCBI does not cover.

Run from pipeline/:
    python build_manual_overrides.py
    python build_manual_overrides.py --input manual_overrides_260731.xlsx --merge
"""

from __future__ import annotations

import argparse
import re
from pathlib import Path

import openpyxl
from Bio.Seq import Seq

from paths import (
    MANUAL_OVERRIDES_FILE, ROOT, gene_names, load_json, load_species_genomes,
    save_json,
)

IN_FILE = ROOT / "manual_overrides_TEMPLATE.xlsx"
SHEET   = "Manual overrides"
HEADER_ROW     = 2
FIRST_DATA_ROW = 3

# Header text (normalised to lowercase alphanumerics) -> field name. Both template
# generations are covered; unlisted headers are ignored.
HEADER_ALIASES = {
    "gene":              "gene",
    "species":           "species",
    "genometosearch":    "genome",
    "status":            "status",
    "proteinaccession":  "accession",
    "proteinsequence":   "protein",
    "cdsrequired":       "cds",
    "cds":               "cds",
    "genomicsequence":   "genomic_seq",
    "reasonnotes":       "reason",
    "reason":            "reason",
    "notes":             "reason",
}
REQUIRED_FIELDS = ("gene", "species", "cds")

# Panel names used before the genes were renamed. Keys are lowercased, so the
# variants that appear in older forms ('AtActin', 'AtACTIN') all resolve.
LEGACY_GENE_NAMES = {
    "atactin":    "ACT1",
    "ataprt":     "APT3",
    "atef1a":     "EF1a",
    "athsp70":    "HSP70",
    "attubulin":  "TUBA1",
    "athistone4": "Histone H4",
    "atgtp_eftu": "EF-Tu",
    "atrpb2":     "RPB2",
    "atubq2":     "RPL40B",
    "atrdr1":     "RDR1",
}

AMINO_ACIDS = set("ACDEFGHIKLMNPQRSTVWYBZJUOX*")


def clean_sequence(value) -> str:
    """Strip whitespace and upper-case a pasted sequence."""
    return re.sub(r"\s+", "", str(value)).upper() if value else ""


def is_protein_sequence(text: str) -> bool:
    """Distinguish a pasted protein from an accession or a free-text note."""
    return len(text) > 20 and set(text) <= AMINO_ACIDS


def translation_error(cds: str, protein: str) -> str | None:
    """None if the CDS translates to `protein`, else a description of the mismatch."""
    if len(cds) % 3:
        return f"CDS length {len(cds)} is not a multiple of 3"
    translated = str(Seq(cds).translate(to_stop=True))
    expected = protein.rstrip("*")
    if translated == expected:
        return None
    return (f"CDS translates to {len(translated)} aa, protein is {len(expected)} aa"
            if len(translated) != len(expected) else
            "CDS translation differs from the pasted protein")


# ── Sheet reading ─────────────────────────────────────────────────────────────

def normalise_header(value) -> str:
    return re.sub(r"[^a-z0-9]", "", str(value or "").lower())


def resolve_columns(sheet) -> dict[str, int]:
    """Map field name -> 0-based column index, from the header row."""
    columns: dict[str, int] = {}
    for index, cell in enumerate(sheet[HEADER_ROW]):
        field = HEADER_ALIASES.get(normalise_header(cell.value))
        if field and field not in columns:
            columns[field] = index

    missing = [f for f in REQUIRED_FIELDS if f not in columns]
    if missing:
        raise SystemExit(
            f"{SHEET!r} row {HEADER_ROW} has no column for: {', '.join(missing)}.\n"
            f"Headers found: "
            f"{', '.join(str(c.value) for c in sheet[HEADER_ROW] if c.value)}"
        )
    return columns


def resolve_gene(label: str, panel: dict[str, str]) -> str | None:
    """Map a sheet gene label onto a current panel name, or None if unknown."""
    key = label.strip().lower()
    return panel.get(key) or LEGACY_GENE_NAMES.get(key)


def resolve_species(label: str, panel: dict[str, str]) -> str | None:
    """Map a sheet species label onto a panel species, or None if unknown.

    Case-insensitive, because a form filled by hand carries 'Vicia Faba' for
    'Vicia faba'. Step 2 looks overrides up by the exact panel spelling, so an
    unnormalised name would be skipped without a word.
    """
    return panel.get(label.strip().lower())


def cell(row: tuple, columns: dict[str, int], field: str):
    index = columns.get(field)
    return row[index] if index is not None and index < len(row) else None


def read_entries(sheet, columns: dict[str, int]
                 ) -> tuple[list[tuple[str, str, dict]], list, list, list, list]:
    """Read the sheet into (entries, skipped, failures, unknown, renamed)."""
    panel = {name.lower(): name for name in gene_names()}
    species_panel = {name.lower(): name for name in load_species_genomes()}

    entries: list[tuple[str, str, dict]] = []
    skipped: list[tuple[str, str]] = []
    failures: list[tuple[str, str, str]] = []
    unknown: list[tuple[int, str, str]] = []
    renamed: list[tuple[int, str, str]] = []

    for offset, row in enumerate(sheet.iter_rows(min_row=FIRST_DATA_ROW,
                                                 values_only=True)):
        row_number = FIRST_DATA_ROW + offset
        label = cell(row, columns, "gene")
        if not row or not label:
            continue

        label    = str(label).strip()
        raw_species = str(cell(row, columns, "species") or "").strip()
        gene     = resolve_gene(label, panel)
        species  = resolve_species(raw_species, species_panel)
        if gene is None:
            unknown.append((row_number, label, raw_species))
            continue
        if species is None:
            unknown.append((row_number, label, f"{raw_species!r} (species not in the panel)"))
            continue
        if species != raw_species:
            renamed.append((row_number, raw_species, species))

        cds = clean_sequence(cell(row, columns, "cds"))
        if not cds:
            skipped.append((gene, species))
            continue

        # The older form has no separate protein column: its "Protein Accession"
        # cell holds the peptide. Detect that so the translation gate still runs.
        protein = clean_sequence(cell(row, columns, "protein"))
        accession = str(cell(row, columns, "accession") or "").strip()
        if not protein and is_protein_sequence(clean_sequence(accession)):
            protein, accession = clean_sequence(accession), ""

        entry = {
            "protein_accession": accession or "manual",
            "cds": cds,
            "cds_length": len(cds),
            "source": "manual",
        }

        if is_protein_sequence(protein):
            entry["protein"] = protein
            error = translation_error(cds, protein)
            if error:
                failures.append((gene, species, error))

        genomic = clean_sequence(cell(row, columns, "genomic_seq"))
        if genomic:
            entry["manual_genomic_sequence"] = genomic
        reason = str(cell(row, columns, "reason") or "").strip()
        if reason:
            entry["manual_modif_reason"] = reason

        entries.append((gene, species, entry))

    return entries, skipped, failures, unknown, renamed


# ── Assembly ──────────────────────────────────────────────────────────────────

def assemble(entries: list[tuple[str, str, dict]], existing: dict
             ) -> tuple[dict, list, list, list]:
    """Combine sheet entries with what is already on file.

    `existing` is {} for a plain rebuild and the current config under --merge.
    Genes come out in panel order; within a gene, entries already on file keep
    their position and new ones are appended.
    """
    overrides: dict[str, dict[str, dict]] = {}
    added, replaced, unchanged = [], [], []

    from_sheet: dict[str, dict[str, dict]] = {}
    for gene, species, entry in entries:
        from_sheet.setdefault(gene, {})[species] = entry

    genes = [g for g in gene_names() if g in existing or g in from_sheet]
    genes += [g for g in (*existing, *from_sheet) if g not in genes]   # off-panel

    kept: list[tuple[str, str]] = []
    for gene in genes:
        previous = existing.get(gene, {})
        current  = from_sheet.get(gene, {})
        merged: dict[str, dict] = {}

        for species, entry in previous.items():
            if species in current:
                new = current[species]
                (unchanged if new.get("cds") == entry.get("cds") else replaced
                 ).append((gene, species))
                merged[species] = new
            else:
                kept.append((gene, species))
                merged[species] = entry

        for species, entry in current.items():
            if species not in merged:
                added.append((gene, species))
                merged[species] = entry

        if merged:
            overrides[gene] = merged

    return overrides, added, replaced, kept


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__.strip().splitlines()[0])
    parser.add_argument("--input", type=Path, default=IN_FILE,
                        help=f"filled form to read (default {IN_FILE.name})")
    parser.add_argument("--merge", action="store_true",
                        help="keep entries already in the config that the form does "
                             "not mention, instead of rewriting the file from the form")
    parser.add_argument("--force", action="store_true",
                        help="write the file even if a CDS fails to translate to its protein")
    args = parser.parse_args()

    in_file = args.input if args.input.is_absolute() else ROOT / args.input
    if not in_file.exists():
        raise SystemExit(f"{in_file} not found — run make_manual_template.py first, "
                         f"then fill it in.")

    workbook = openpyxl.load_workbook(in_file, data_only=True)
    if SHEET not in workbook.sheetnames:
        raise SystemExit(f"{in_file.name} has no {SHEET!r} sheet "
                         f"(found: {', '.join(workbook.sheetnames)})")
    sheet = workbook[SHEET]

    columns = resolve_columns(sheet)
    entries, skipped, failures, unknown, renamed = read_entries(sheet, columns)

    if renamed:
        print(f"{len(renamed)} species name(s) normalised to the panel spelling:")
        for row_number, was, now in renamed:
            print(f"    row {row_number}: {was!r} -> {now!r}")

    if unknown:
        print(f"{len(unknown)} row(s) name a gene or species that is not in the panel:")
        for row_number, label, species in unknown:
            print(f"    row {row_number}: {label!r} / {species}")
        raise SystemExit(f"\nNothing written. Panel: {', '.join(gene_names())}")

    existing = load_json(MANUAL_OVERRIDES_FILE) if args.merge else {}
    overrides, added, replaced, kept = assemble(entries, existing)

    total = sum(len(v) for v in overrides.values())
    validated = sum(1 for _, _, e in entries if "protein" in e) - len(failures)
    print(f"{len(entries)} rows read from {in_file.name}")
    print(f"  {validated} validated (CDS translates to the pasted protein)")
    if skipped:
        print(f"  {len(skipped)} row(s) skipped (no CDS):")
        for gene, species in skipped:
            print(f"    {gene} / {species}")
    for label, group in (("added", added), ("replaced", replaced), ("kept", kept)):
        if group:
            print(f"  {len(group)} {label}:")
            for gene, species in group:
                print(f"    {gene} / {species}")
    if failures:
        print(f"  {len(failures)} FAILED validation:")
        for gene, species, error in failures:
            print(f"    {gene} / {species}: {error}")
        if not args.force:
            raise SystemExit("\nNothing written. Fix the rows above, or pass --force.")

    save_json(MANUAL_OVERRIDES_FILE, overrides)
    print(f"\nWrote {MANUAL_OVERRIDES_FILE} — {total} entries "
          f"across {len(overrides)} genes")


if __name__ == "__main__":
    main()
